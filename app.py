
from flask import Flask, render_template, request, redirect, url_for, send_file, session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sqlite3
from pathlib import Path
from datetime import datetime
import io, os

BASE = Path(__file__).resolve().parent
DB = BASE / "data" / "guests.db"
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-this-secret-key")

def db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    DB.parent.mkdir(exist_ok=True)
    conn = db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS guests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            attendance TEXT NOT NULL,
            companion TEXT NOT NULL,
            alcohol TEXT,
            transfer TEXT,
            allergies TEXT,
            children TEXT,
            comments TEXT,
            created_at TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()

@app.route("/")
def index():
    return render_template("index.html")

@app.post("/rsvp")
def rsvp():
    fields = ["name", "attendance", "companion", "alcohol", "transfer", "allergies", "children", "comments"]
    data = {f: request.form.get(f, "").strip() for f in fields}
    if not data["name"] or not data["attendance"] or not data["companion"]:
        return redirect(url_for("index", error="Заполните обязательные поля"))
    conn = db()
    conn.execute("""
        INSERT INTO guests
        (name, attendance, companion, alcohol, transfer, allergies, children, comments, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (*[data[f] for f in fields], datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()
    return redirect(url_for("index", sent="1"))

@app.route("/admin")
def admin():
    # Для первого запуска: /admin?key=change-me
    if request.args.get("key") != os.environ.get("ADMIN_KEY", "change-me"):
        return "Доступ запрещён. Укажите правильный ADMIN_KEY.", 403
    conn = db()
    guests = conn.execute("SELECT * FROM guests ORDER BY id DESC").fetchall()
    conn.close()
    return render_template("admin.html", guests=guests)

@app.route("/admin/export.xlsx")
def export_xlsx():
    if request.args.get("key") != os.environ.get("ADMIN_KEY", "change-me"):
        return "Доступ запрещён.", 403
    conn = db()
    rows = conn.execute("SELECT * FROM guests ORDER BY id ASC").fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Гости"
    headers = ["№", "Имя и фамилия", "Присутствие", "Формат", "Алкоголь",
               "Трансфер", "Аллергии / питание", "Дети", "Комментарий", "Дата ответа"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="5F6848")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, r in enumerate(rows, 1):
        ws.append([i, r["name"], r["attendance"], r["companion"], r["alcohol"],
                   r["transfer"], r["allergies"], r["children"], r["comments"], r["created_at"]])

    widths = [6, 28, 18, 18, 28, 15, 30, 12, 40, 22]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True,
                     download_name="Гости_Виктор_Валерия.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
